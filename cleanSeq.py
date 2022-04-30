import sys
import time
import os
from bs4 import BeautifulSoup
import requests
import random
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl import Workbook
from Bio import SeqIO
from Bio import Entrez
from reportlab.platypus import SimpleDocTemplate, Paragraph,Table,Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import pandas as pd

def mutationFile(referenceGenome,vcf):
    ref = open(referenceGenome,'r')
    ref_lines = ref.readlines();
    ref_lines = ref_lines[1:]
    ref_str = ''.join(ref_lines)
    ref_str = ref_str.replace('\n','')
    ref_str = '0' + ref_str
    ref_list = list(ref_str)
    alt = open(vcf,'r')
    alt_lines = alt.readlines();
    mutation_info = pd.DataFrame(columns=['pos','ref','alt'])
    for line in alt_lines:
        if line[0] != '#':
            mutation_info = mutation_info.append({'pos':int(line.split('\t')[1]),'alt':line.split('\t')[4],'ref':line.split('\t')[3]},ignore_index = True)
    alt_dict = dict(zip(range(len(ref_list)),ref_list))
    for i in range(len(mutation_info['pos'])):
        temp_pos = mutation_info['pos'][i]
        temp_ref = mutation_info['ref'][i]
        temp_alt = mutation_info['alt'][i]
        for j in range(len(temp_ref)):
            if j == 0:
                alt_dict[temp_pos] = temp_alt
            else:    
                alt_dict[temp_pos+j] = '0'
    return mutation_info,ref_list,alt_dict

def Mutquery(referenceGenome,vcf):
    mutation_info,ref_list,alt_dict = mutationFile(referenceGenome,vcf)
    pos_result = []
    i = 0
    while i < len(mutation_info['pos']):
        j = 0
        start_pos = mutation_info['pos'][i]
        while i + j + 1 < len(mutation_info['pos']):
            if mutation_info['pos'][i + j + 1] - mutation_info['pos'][i + j] < 8:
                if(i + j + 1 == len(mutation_info['pos'])-1):
                    end_pos = mutation_info['pos'][i + j + 1]
                    pos_result.append([start_pos - 25, end_pos + 50,start_pos, end_pos])
                    j += 10
                    break;
                j += 1
                continue;
            else:
                j += 1
                end_pos = mutation_info['pos'][i + j - 1]
                pos_result.append([start_pos - 25, end_pos + 50,start_pos, end_pos])
                break;
        i = i + j
        if(i >= len(mutation_info['pos'])-1):
           break
    mut_result = []
    mut_str = ""
    for i in range(len(pos_result)):
        start_index = pos_result[i][0]
        end_index = pos_result[i][1]
        if pos_result[i][2] != pos_result[i][3]:
            for j in range(start_index, end_index+100):
                mut_str += alt_dict.get(j)
                mut_str = mut_str.replace('0','')
                if len(mut_str) > 50:
                    mut_str = mut_str[0:50]
                    end_index = j 
                    break
            mut_result.append([start_index,end_index,mut_str,pos_result[i][2],pos_result[i][3]])
            mut_str = ""
            start_index = 0
            end_index = 0
    file = open("output/Mutquery.fasta", 'w')
    for i in range(len(mutation_info['pos'])):
        temp_pos = mutation_info['pos'][i]
        temp_alt = mutation_info['alt'][i]
        temp_ref = mutation_info['ref'][i]
        mutseq = "".join(ref_list[temp_pos-25:temp_pos]) + temp_alt + "".join(ref_list[temp_pos + len(temp_ref) : temp_pos + 50])
        mutseq = mutseq[:50]
        pos = ">" + str(temp_pos - 25) + "." + str(temp_pos + 25) + "." + str(temp_pos) + "." + str(temp_pos)
        file.write(str(pos) + '\n')
        file.write(mutseq + '\n')
    for i in range(len(mut_result)):
        fasta_index = ">"+str(mut_result[i][0])+"."+str(mut_result[i][1])
        fasta_index2 = str(mut_result[i][3]) + "." + str(mut_result[i][4])
        fasta_index = fasta_index + "." + fasta_index2
        file.writelines(fasta_index + '\n')
        file.writelines(str(mut_result[i][2]) + '\n')
    file.close()    

def save_optimatresult(dict1,filename):
    if isinstance(dict1, str):
        dict1 = eval(dict1)
    with open(filename, 'w', encoding='utf-8') as f:
        for key,value in dict1.items():
            str_ = str(key)
            f.write(str_)
            f.write('\n')
            for i in range(len(value)):            
                if int(eval(value[i][8])) < int(eval(value[i][9])):
                    f.write(str(value[i]))
                    f.write('\n')

def fastqtofasta(fastq,fasta):
    records = SeqIO.parse(fastq, "fastq")
    SeqIO.write(records, fasta, "fasta")
    
def fastaTofastq(fasta,fastq,mapp,out):
    commond = "grep -E '@|F' " + fastq + " > " + mapp
    os.system(commond)
    file = open(mapp,'r')
    dictmapp = {}
    i = 0
    key = ""
    value = ""
    while True:
        line = file.readline()
        if not line:
            break
        if i % 2 == 0:
            key = line.split('@')[1].split(' ')[0]
        elif i % 2 == 1:
            value = line
            dictmapp[key] = value
        i = i + 1    
    record = SeqIO.index(fasta,'fasta')   
    fout = open(out,'w')
    for key in record.keys():
        first = "@"+str(key)+'\n'
        second = str(record[key].seq)+'\n'
        third = "+\n"
        fourth = dictmapp[str(key)]
        fout.write(first)
        fout.write(second)
        fout.write(third)
        fout.write(fourth)

def deal_outfmt7(filename):
    with open(filename,'r') as f1:
        text = f1.readlines()
    desc_list = []
    data_list = []
    for row in text:
        if row[0] == "#":
            if row[0:3] == '# B' or row[0:3] == '# D' or row[0:3] == '# F':
                pass
            else:
                desc_list.append(row)
        else:
            data_list.append(row)
    index = 0
    desc_dict = {}
    while(index < len(desc_list)):
        key = desc_list[index]
        key = (key.split(' ')[2]).strip('\n')
        value = desc_list[index+1]
        value = int(value.split(' ')[1])
        index += 2
        if key in desc_dict:
            desc_dict[key] += value
        else:
            desc_dict[key] = value
    data_list_new = []
    for row in data_list:
        data_list_new.append(row.split('\t'))
    result_dict = {}
    index = 0
    for key,value in desc_dict.items():
        result_dict[key] = data_list_new[index:index+value]
        index = index + value
    return result_dict

def optimtal_blast_result(inputfile,outputfile):
    file_dict = deal_outfmt7(inputfile)
    result_dict = {}
    value_list = []
    for key,values in file_dict.items():
        for value in values:
            cond1 = int(key.split('.')[0]) + int(value[6])
            cond2 = int(key.split('.')[0]) + int(value[7])
            if int(value[4]) < 1 and int(value[5]) < 1:
                if cond1 < int(key.split('.')[2]):
                    if cond2 > int(key.split('.')[3]):
                        value_list.append(value) 
        result_dict[key] = value_list
        value_list = []
    save_optimatresult(result_dict, outputfile)

def extract_reads(rawdata,file):
    mutation_reads_index = {}
    values = []
    with open(file) as transcripts:
        for line in transcripts:
            if line in ['\n','\r\n']:
                continue
            if len(line.split(':')) > 4:
                key = line.strip().split(',')[0].split('\'')[1]
                if key in mutation_reads_index:
                    values.append(line.strip().split(',')[1].split('\'')[1])
                    mutation_reads_index[key] = values
                else:
                    values = []
                    values.append(line.strip().split(',')[1].split('\'')[1])
                    mutation_reads_index[key] = values
    return mutation_reads_index

def Kmer(referenceGenome,outfile):
    ref = open(referenceGenome,'r')
    ref_lines = ref.readlines();
    ref_lines = ref_lines[1:]
    ref_str = ''.join(ref_lines)
    ref_str = ref_str.replace('\n','')
    ref_str = '0' + ref_str
    ref_list = list(ref_str)
    ref_dict = dict(zip(range(len(ref_list)),ref_list))
    ref_str = ""
    ref_result = []
    strPos = 0
    endPos = 0
    count = 0
    for key,value in ref_dict.items():
        if count == 150:
            strPos = key - count
            endPos = key
            ref_str = ref_str + str(value)
            ref_result.append([strPos,endPos,ref_str])
            count = 0
            ref_str = ""
        else:
            ref_str = ref_str + str(value)
            count += 1
    file = open(outfile, 'w')
    for i in range(len(ref_result)):
        fasta_index = ">"+str(ref_result[i][0])+"_"+str(ref_result[i][1])
        file.writelines(fasta_index)
        file.write('\n')
        if i == 0:
            file.writelines(str(ref_result[i][2])[1:])
        else:
            file.writelines(str(ref_result[i][2]))
        file.write('\n')
    file.close()

def processAlignFile(alnFilePath):
    valueStr = ""
    values = []
    key = []
    keyStr = ""
    file = open(alnFilePath,'r')
    f = file.readlines()
    f.append('>')
    for line in f:
        if line.split('>')[0] == "":
            key.append(keyStr.strip('\n'))
            keyStr = line
            values.append(list(valueStr))
            valueStr = ""
        else:
            valueStr += line.strip('\n')
    Mseq = values[1]
    startPos = 0
    for i in Mseq:
        if i == '-':
            startPos += 1
        else:
            break
    mutantBase = Mseq[startPos+25]
    
    df = pd.DataFrame(values[1:],index = key[1:])
    base = {}
    for i in range(df.shape[1]):
        temp = {}
        countA  = 0
        countT  = 0
        countC  = 0
        countG  = 0
        for s in df[i].values:
            if s == 'A':
                countA += 1
            elif s == 'T':
                countT += 1
            elif s == 'C':
                countC += 1
            elif s == 'G':
                countG += 1
        temp['A'] = countA
        temp['T'] = countT
        temp['C'] = countC
        temp['G'] = countG
        base[i] = temp
        temp = {}
    return startPos,base,mutantBase

def mergeSameNameFile(inputPath1,inputPath2,outpath):
    def merge(file1,file2):
        if not os.path.exists(outpath):
            os.mkdir(outpath)
        filename = outpath + "/" + file1.split('/')[-1]
        file = open(filename,'w')
        for line in open(file1):
            file.writelines(line)
        file.write('\n')
        for line in open(file2):
            file.writelines(line)
        file.write('\n')
        file.close()
    list1 = os.listdir(inputPath1)
    list2 = os.listdir(inputPath2)
    for l in list1:
        if l in list2:
            file1 = inputPath1 + "/" + l
            file2  = inputPath2 + "/" + l
            merge(file1,file2)

def multipeBase(dictBase):
    count = 0
    base = []
    base.append(dictBase['A'])
    base.append(dictBase['C'])
    base.append(dictBase['T'])
    base.append(dictBase['G'])
    count = sum(ℹ > 0 for i in base)
    return count
        
def coverage(filePath):
    filename = os.listdir(filePath)
    namecount = 0
    os.system('touch output/extract_result/misalignment.txt')
    fout = open("output/extract_result/misalignment.txt",'w')
    for name in filename:
        if name[-3:] == 'aln':
            name = filePath + name
            mutPos,mutdict,mutantBase = processAlignFile(name)
            countmisalignment = 0
            for i in range(mutPos,mutPos + 50):
                flag = multipeBase(mutdict[i])
                if flag > 1:
                    countmisalignment += 1
            if countmisalignment > 20:
                namecount = namecount + 1
                fout.write(name)
                fout.write('\n')
    fout.write('total have ' + str(namecount) + ' misalignment file, please check manually.')
    fout.write('\n')
    fout.close()

def clustalo(filePath):
    filename = os.listdir(filePath)
    for name in filename: 
        outname = name.split('f')[0] + "aln"
        name = filePath + name
        outname = filePath + outname
        command = "./clustalo -i " + name + " -o " + outname
        os.system(command)

def delet():
    if os.path.exists('output/blast1.txt'):os.remove('output/blast1.txt')
    if os.path.exists('output/blast2.txt'):os.remove('output/blast2.txt')
    if os.path.exists('output/contamination.txt'):os.remove('output/contamination.txt')
    if os.path.exists('output/fblast.txt'):os.remove('output/fblast.txt')
    if os.path.exists('output/fextract.fasta'):os.remove('output/fextract.fasta')
    if os.path.exists('output/Mutquery.fasta'):os.remove('output/Mutquery.fasta')
    if os.path.exists('output/opt_blast1.txt'):os.remove('output/opt_blast1.txt')
    if os.path.exists('output/opt_blast2.txt'):os.remove('output/opt_blast2.txt')
    if os.path.exists('output/PKmer.fasta'):os.remove('output/PKmer.fasta')
    if os.path.exists('output/r1.fasta'):os.remove('output/r1.fasta')
    if os.path.exists('output/r1.txt'):os.remove('output/r1.txt')
    if os.path.exists('output/random.fasta'):os.remove('output/random.fasta')
    if os.path.exists('output/raw1.fasta'):os.remove('output/raw1.fasta')
    if os.path.exists('output/raw2.fasta'):os.remove('output/raw2.fasta')
    if os.path.exists('output/RKmer.fasta'):os.remove('output/RKmer.fasta')
    if os.path.exists('output/sblast.txt'):os.remove('output/sblast.txt')
    if os.path.exists('output/GATK/R.bam'):os.remove('output/GATK/R.bam')
    if os.path.exists('output/GATK/R.sam'):os.remove('output/GATK/R.sam')
    if os.path.exists('output/GATK/R.filter.INDEL.vcf'):os.remove('output/GATK/R.filter.INDEL.vcf')
    if os.path.exists('output/GATK/R.filter.INDEL.vcf.idx'):os.remove('output/GATK/R.filter.INDEL.vcf.idx')
    if os.path.exists('output/GATK/R.filter.SNP.vcf'):os.remove('output/GATK/R.filter.SNP.vcf')
    if os.path.exists('output/GATK/R.filter.SNP.vcf.idx'):os.remove('output/GATK/R.filter.SNP.vcf.idx')
    if os.path.exists('output/GATK/R.filter.vcf.idx'):os.remove('output/GATK/R.filter.vcf.idx')
    if os.path.exists('output/GATK/R.gvcf'):os.remove('output/GATK/R.gvcf')
    if os.path.exists('output/GATK/R.gvcf.idx'):os.remove('output/GATK/R.gvcf.idx')
    if os.path.exists('output/GATK/R.INDEL.vcf'):os.remove('output/GATK/R.INDEL.vcf')
    if os.path.exists('output/GATK/R.INDEL.vcf.idx'):os.remove('output/GATK/R.INDEL.vcf.idx')
    if os.path.exists('output/GATK/R.SNP.vcf'):os.remove('output/GATK/R.SNP.vcf')
    if os.path.exists('output/GATK/R.SNP.vcf.idx'):os.remove('output/GATK/R.SNP.vcf.idx')
    if os.path.exists('output/GATK/R.sorted.bam.bai'):os.remove('output/GATK/R.sorted.bam.bai')
    if os.path.exists('output/GATK/R.sorted.markdup.bam'):os.remove('output/GATK/R.sorted.markdup.bam')
    if os.path.exists('output/GATK/R.sorted.markdup.bam.bai'):os.remove('output/GATK/R.sorted.markdup.bam.bai')
    if os.path.exists('output/GATK/R.sorted.markdup_metrix.txt'):os.remove('output/GATK/R.sorted.markdup_metrix.txt')
    if os.path.exists('output/GATK/R.vcf'):os.remove('output/GATK/R.vcf')
    if os.path.exists('output/GATK/R.vcf.idx'):os.remove('output/GATK/R.vcf.idx')
    if os.path.exists('output/clean.hist'):os.remove('output/clean.hist')
    if os.path.exists('output/clean.histogram'):os.remove('output/clean.histogram')
    if os.path.exists('output/clean.notCombined_1.fastq'):os.remove('output/clean.notCombined_1.fastq')
    if os.path.exists('output/clean.notCombined_2.fastq'):os.remove('output/clean.notCombined_2.fastq')
    if os.path.exists('output/cleanup1.mapp'):os.remove('output/cleanup1.mapp')
    if os.path.exists('output/cleanup2.mapp'):os.remove('output/cleanup2.mapp')
    if os.path.exists('output/log'):os.remove('output/log')
    if os.path.exists('output/out.trim1.fastq'):os.remove('output/out.trim1.fastq')
    if os.path.exists('output/out.trim2.fastq'):os.remove('output/out.trim2.fastq')

def random10000(rawdata):
    records_index = SeqIO.index(rawdata,'fasta')
    randomKey = random.sample(records_index.keys(), 10000)
    fout = open("output/random.fasta",'w')
    for key in randomKey:
        fout.write(">" + key + "\n")
        fout.write(str(records_index[key].seq) + "\n")

def condetect_blast(dbPath):
    command = "./blastn -query output/random.fasta -db " + dbPath + " -out output/contamination.txt -num_threads 32 -evalue 0.00000000001 -outfmt 6 -max_target_seqs 5"
    os.system(command) 
    file2 = open("output/contamination.txt",'r')
    lines = file2.readlines()
    result_dict = {}
    for row in lines:
        if row.isspace():
            continue
        key = row.split('\t')[0]
        if key in result_dict.keys():
            value = row.split('\t')[1:6]
            if float(value[3]) <= 1 and float(value[4]) <= 1:
                if float(value[1]) > float(result_dict[key][1]):
                    result_dict[key] = value
            else:
                continue
        else:
            value = row.split('\t')[1:6]
            result_dict[key] = value
    return result_dict

def getHTMLText(url):
    try:
        r = requests.get(url, timeout = 30)
        r.raise_for_status()
        r.encoding = r.apparent_encoding
        return r.text
    except:
        return ""

def parse_page(html):
    soup = BeautifulSoup(html, "html.parser")
    return soup.title

def originam(info):
    species = {}
    species2 = {}
    total_sum = 0
    for line in info.values():
        num = int(line.split(':')[1])
        total_sum += num
        if 'None' not in line:
            key = ' '.join(line.split(':')[2].split('<title>')[1].split(',')[0].split(' ')[:2])
            if key in species.keys():
                species[key] = species[key] + num
            else:
                species[key] = num
    for key in species.keys():
        for value in info.values():
            if key in value:
                species2[key + ':' +value.split(':')[0]] = species[key]
                break
    matchedCount = total_sum
    speciesDict = sorted(species2.items(), key=lambda x: x[1], reverse=True)
    return matchedCount,speciesDict

def PlotStat(matchedCount,speciesDict):
    dict1 = {}
    dict2 = {}
    TopTensum = 0
    for item in speciesDict[:10]:
        key = item[0].split(':')[0]
        value = item[1] / 10000
        dict1[key] = value * 100
        dict2[key] = item[1]
        TopTensum += item[1]

    others = matchedCount - TopTensum
    Noresult = 10000 - matchedCount
    dict1['others'] = others / matchedCount *100
    dict1['Unknow'] = Noresult / 10000 * 100
    dict2['others'] = others
    dict2['Unknow'] = Noresult

    fig, ax1 = plt.subplots(figsize=(6,6))
    xticks = range(0,len(dict1.keys()))
    xlabels = dict1.keys()
    ax1.set_xticks(xticks)
    ax1.set_xticklabels(xlabels, rotation =90)
    ax1.set_ylabel("Matched ratio (%)", fontsize = 15)
    ax1.set_title("Contamination Report", fontsize = 20)
    ax1.set_xlabel('Species Name (Top ten)', fontsize = 15)
    plt.bar(dict1.keys(),dict1.values())  
    ax2 = ax1.twinx()
    plt.bar(dict2.keys(),dict2.values())
    ax2.set_ylabel('Num of matched Reads',fontsize = 15)
    plt.savefig('output/Species.png', dpi=400, bbox_inches='tight')
    
def speciesExcel(speciesDict):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'species'
    wb.save("output/Species.xlsx")
    excel = load_workbook('output/Species.xlsx')
    table = excel['species']
    row_index = 2
    col_index = 1
    table.cell(row = 1, column = 1,value = "Species Name")
    table.cell(row = 1, column = 2,value = "num of Matched reads")
    table.cell(row = 1, column = 3,value = "accessionID")
    for item in speciesDict:
        table.cell(row = row_index,column = col_index, value = item[0].split(':')[0])
        table.cell(row = row_index,column = col_index + 1, value = item[1])
        table.cell(row = row_index,column = col_index + 2, value = item[0].split(':')[1])
        row_index += 1
    excel.save('output/Species.xlsx')

def similary(r1,r2):
    Kmer(r1, 'output/r1.fasta')
    query_path = 'output/r1.fasta'
    makeblastdb = "./makeblastdb -dbtype nucl -in " + r2 + " -input_type fasta -out output/db/r2.blastdb -logfile output/db/log.txt"
    blast = "./blastn -query " + query_path + " -db output/db/r2.blastdb -out output/r1.txt -num_threads 20 -evalue 0.000001 -outfmt 7 "
    os.system(makeblastdb)
    os.system(blast)
    dict1 = deal_outfmt7('output/r1.txt')
    count = len(dict1)
    countGreat90 = 0
    for v in dict1.values():
        if len(v) == 0:
            continue
        if float(v[0][2]) >= 90:
            countGreat90 += 1
    ani = countGreat90 / count * 100
    #print(count,countGreat90,ani)
    return ani

def getContaminateRef(speciesDict,reference):
    contaminateKey = []
    for item in speciesDict[:11]:
        key = item[0].split(':')[1]
        value = item[1] / 10000 * 100
        if value > 10:
            contaminateKey.append(key)

    Entrez.email = "bingxiao_m@163.com"
    contaminationReference = []
    for key in contaminateKey:
        f = Entrez.efetch(db="nucleotide",id=key,rettype='fasta')
        seq = SeqIO.read(f,'fasta')
        contaminateRef = 'output/' + key + '.fasta'
        os.system("touch " + contaminateRef)
        fw = open(contaminateRef,'w')
        SeqIO.write(seq,fw,'fasta')
        fw.close()
        os.getcwd()
        ani = similary(reference, contaminateRef)
        if ani < 80:
            contaminationReference.append(contaminateRef)
    return contaminationReference

def contamiateDetection(rawdata,dbPath,reference):
    random10000(rawdata)
    result_dict = condetect_blast(dbPath)
    speciesRef = {}
    for v in result_dict.values():
        if v[0] in speciesRef.keys():
            speciesRef[v[0]] = speciesRef[v[0]] + 1
        else:
            speciesRef[v[0]] = 1
    info = {}
    for key in speciesRef.keys():
        url = 'https://www.ncbi.nlm.nih.gov/nuccore/' + key
        try:
            html = getHTMLText(url)
            originamStr = parse_page(html)
            info[key] = str(key) +':'+  str(speciesRef[key]) + ":" + str(originamStr) 
        except:
            print("wrong")
    matchedCount,speciesDict = originam(info)
    PlotStat(matchedCount,speciesDict)
    speciesExcel(speciesDict)
    conref = getContaminateRef(speciesDict,reference)
    return conref

def firstBlast(referenceGenome,rawdata1,rawdata2):
    Kmer(referenceGenome,"output/RKmer.fasta")
    query_path = "output/RKmer.fasta"
    makeblastdb = "./makeblastdb -dbtype nucl -in " + rawdata1 + " -input_type fasta -out output/db/clean.blastdb -logfile output/db/log.txt" 
    blast_cmd = "./blastn -query " + query_path + " -db output/db/clean.blastdb -out output/fblast.txt -num_threads 20 -evalue 0.000001 -outfmt 7"
    os.system(makeblastdb)
    os.system(blast_cmd)

def secondBlast(PreferenceGenome,data1):
    Kmer(PreferenceGenome,"output/PKmer.fasta")
    query_path = "output/PKmer.fasta"
    makeblastdb = "./makeblastdb -dbtype nucl -in " + data1 + " -input_type fasta -out output/db/clean.blastdb -logfile output/db/log.txt"
    blast_cmd = "./blastn -query " + query_path + " -db output/db/clean.blastdb -out output/sblast.txt -num_threads 20 -evalue 0.00000000001 -outfmt 7"
    os.system(makeblastdb)
    os.system(blast_cmd)
    
def firstBlast_extract(file,outfile,rawdata):
    f1 = deal_outfmt7(file)
    f1_l = []
    for value in f1.values():
        for v in value:
            f1_l.append(v[1])
    f1_l = list(set(f1_l))
    fout = open(outfile,'w')  
    records_index = SeqIO.index(rawdata,'fasta')
    for item in f1_l:
        fout.write(">" + str(records_index[str(item)].description) + "\n")
        fout.write(str(records_index[item].seq) + "\n")
    fout.close()

def cleanup_optimtal_blast(inputfile,similarity):
    file_dict = deal_outfmt7(inputfile)
    value_list = []
    for key,values in file_dict.items():
        for value in values:
            if float(value[2]) > float(similarity): 
                value_list.append(value[1]) 
    return value_list
    
def secondBlast_extract(fb,sb,out1,out2,r1,r2,similarity):
    M1 = deal_outfmt7(fb)
    Mlist = []
    Plist = []
    for value in M1.values():
        for v in value:
            Mlist.append(v[1]) 
    Pdict = {} 
    Plist = cleanup_optimtal_blast(sb,similarity)   
    Mlist = list(set(Mlist))
    Plist = list(set(Plist))
    for i in Plist:
        Pdict[i] = 1
    fout = open(out1,'w')  
    records_index = SeqIO.index(r1,'fasta')
    count = 0
    for item in Mlist:
        if item in Pdict.keys():
            count = count + 1
            continue
        else:
            fout.write(">" + str(records_index[str(item)].description) + "\n")
            fout.write(str(records_index[item].seq) + "\n")
    fout.close()
    fout = open(out2,'w')  
    records_index = SeqIO.index(r2,'fasta')
    for item in Mlist:
        if item in Pdict.keys():
            continue
        else:
            fout.write(">" + str(records_index[str(item)].description) + "\n")
            fout.write(str(records_index[item].seq) + "\n")
    fout.close()

def cleanup(referenceGenome,ContamreferenceGenome,rawdata1,rawdata2,similarity):
    firstBlast(referenceGenome, rawdata1, rawdata2)
    firstBlast_extract("output/fblast.txt","output/fextract.fasta",rawdata1)
    secondBlast(ContamreferenceGenome,"output/fextract.fasta")
    secondBlast_extract("output/fblast.txt","output/sblast.txt","output/cleanupReads1.fasta","output/cleanupReads2.fasta",rawdata1,rawdata2,similarity)

def pairEnd_blast(referenceGenome,vcf,rawdata1,rawdata2):
    Mutquery(referenceGenome, vcf)
    query_path = "output//Mutquery.fasta"
    makeblastdb1 = "./makeblastdb -dbtype nucl -in " + rawdata1 + " -input_type fasta -out output/db/vcf1.blastdb -logfile output/db/log.txt"
    blast_cmd1 = "./blastn -query " + query_path + " -db output/db/vcf1.blastdb -out output/blast1.txt -num_threads 32 -word_size 11 -evalue 0.000001 -outfmt 7"
    
    makeblastdb2 = "./makeblastdb -dbtype nucl -in " + rawdata2 + " -input_type fasta -out output/db/vcf2.blastdb -logfile output/db/log.txt"
    blast_cmd2 = "./blastn -query " + query_path + " -db output/db/vcf2.blastdb -out output/blast2.txt -num_threads 32 -word_size 11 -evalue 0.000001 -outfmt 7"
    os.system(makeblastdb1)
    os.system(makeblastdb2)
    os.system(blast_cmd1)
    os.system(blast_cmd2)

def pairEnd_extract(outputPath,rawdata,file,flag):
    if not os.path.exists(outputPath):
        os.makedirs(outputPath)
    mutation_reads_index = extract_reads(rawdata, file)
    records_query = SeqIO.index('output/Mutquery.fasta','fasta')
    records_index = SeqIO.index(rawdata,'fasta')
    if len(mutation_reads_index) > 0:
        if flag == True:
            for key,value in mutation_reads_index.items():
                path = outputPath + key + '.fasta'
                fout = open(path,'w')
                fout.write(">" + key + "\n")
                fout.write(str(records_query[key].seq)+ "\n")
                value = list(set(value))
                for i in range(len(value)):
                    fout.write(">" + value[i] + "\n")
                    fout.write(str(records_index[value[i]].seq) + "\n")
            fout.close()
        else:
            for key,value in mutation_reads_index.items():
                path = outputPath + key + '.fasta'
                fout = open(path,'w')
                value = list(set(value))
                for i in range(len(value)):
                    fout.write(">" + value[i] + "\n")
                    fout.write(str(records_index[value[i]].seq) + "\n")
            fout.close()
  
def pairEndBlast(referenceGenome,vcf,rawdata1,rawdata2):
    pairEnd_blast(referenceGenome,vcf,rawdata1,rawdata2)
    optimtal_blast_result("output/blast1.txt", "output/opt_blast1.txt")
    optimtal_blast_result("output/blast2.txt", "output/opt_blast2.txt")
    pairEnd_extract("output/extract_result/first/",rawdata1,"output/opt_blast1.txt",True)
    pairEnd_extract("output/extract_result/second/",rawdata2,"output/opt_blast2.txt",False)
    mergeSameNameFile("output/extract_result/first", "output/extract_result/second", "output/extract_result/")
    clustalo("output/extract_result/")
    coverage("output/extract_result/")

def SingleEnd_blast(referenceGenome,vcf,rawdata1):
    Mutquery(referenceGenome, vcf)
    query_path = "output//Mutquery.fasta"
    makeblastdb1 = "./makeblastdb -dbtype nucl -in " + rawdata1 + " -input_type fasta -out output/db/singleEndvcf1.blastdb -logfile output/db/log.txt"
    blast_cmd1 = "./blastn -query " + query_path + " -db output/db/singleEndvcf1.blastdb -out output/blast1.txt -num_threads 32 -word_size 11 -evalue 0.000001 -outfmt 7"
    os.system(makeblastdb1)
    os.system(blast_cmd1)

def SingleEnd_extract(outputPath,rawdata,file):
    if not os.path.exists(outputPath):
        os.makedirs(outputPath)
    mutation_reads_index = extract_reads(rawdata, file)
    records_query = SeqIO.index('output/Mutquery.fasta','fasta')
    records_index = SeqIO.index(rawdata,'fasta')
    if len(mutation_reads_index) > 0:
        for key,value in mutation_reads_index.items():
            path = outputPath + key + '.fasta'
            fout = open(path,'w')
            fout.write(">" + key + "\n")
            fout.write(str(records_query[key].seq)+ "\n")
            value = list(set(value))
            for i in range(len(value)):
                fout.write(">" + value[i] + "\n")
                fout.write(str(records_index[value[i]].seq) + "\n")
        fout.close()
        
def singleEndBlast(referenceGenome,vcf,rawdata1):
    SingleEnd_blast(referenceGenome,vcf,rawdata1)
    optimtal_blast_result("output/blast1.txt", "output/opt_blast1.txt")
    SingleEnd_extract("output/extract_result/",rawdata1,"output/opt_blast1.txt")
    clustalo("output/extract_result/")
    coverage("output/extract_result/")

def stat_notcon(reference,raw1):
    recordsRef = SeqIO.index(reference, 'fasta')
    records1 = SeqIO.index(raw1,'fasta')
    coverage1 = 0
    count1 = 0
    basecount = 0
    NumVariants = 0
    NumVerifed = 0
    NumMisalign = 0
    for key in recordsRef.keys():
        countref = len(str(recordsRef[key].seq))
    for key in records1.keys():
        basecount += len(str(records1[key].seq))
        count1 += 1
    coverage1 = basecount/countref
    fvcf = open('output/GATK/R.filter.xls','r')
    fmisalign = open('output/extract_result/misalignment.txt','r')
    filename = os.listdir('output/extract_result/')
    for name in filename: 
        if name[-3:] == 'aln':
            NumVerifed += 1
    NumVariants = len(fvcf.readlines()) - 1
    NumMisalign = len(fmisalign.readlines()) - 1
    return count1,coverage1,NumVariants,NumVerifed,NumMisalign

def stat_con(reference,raw1,cleanupreads1):
    recordsRef = SeqIO.index(reference, 'fasta')
    records1 = SeqIO.index(raw1,'fasta')
    records2 = SeqIO.index(cleanupreads1,'fasta')
    coverage1 = 0
    coverage2 = 0
    count1 = 0
    count2 = 0
    basecount = 0
    NumVariants = 0
    NumVerifed = 0
    NumMisalign = 0
    for key in recordsRef.keys():
        countref = len(str(recordsRef[key].seq))
    for key in records1.keys():
        basecount += len(str(records1[key].seq))
        count1 += 1
    coverage1 = basecount/countref
    basecount = 0
    for key in records2.keys():
        basecount += len(str(records1[key].seq))
        count2 += 1
    coverage2 = basecount/countref
    fvcf = open('output/GATK/R.filter.xls','r')
    fmisalign = open('output/extract_result/misalignment.txt','r')
    filename = os.listdir('output/extract_result/')
    for name in filename: 
        if name[-3:] == 'aln':
            NumVerifed += 1
    NumVariants = len(fvcf.readlines()) - 1
    NumMisalign = len(fmisalign.readlines()) - 1
    return count1,coverage1,count2,coverage2,NumVariants,NumVerifed,NumMisalign
  
def contaminationReport(reference,rawfastq1,rawfastq2,conref):
    doc = SimpleDocTemplate("output/Report.pdf")
    styles = getSampleStyleSheet()
    
    p = styles['Normal']
    p.fontSize = 10
    p.leftIndent = -20
    p.spaceAfter = 15
    p.spaceBefore = 15
    
    smallT = styles['BodyText']
    smallT.fontSize = 15
    smallT.spaceAfter = 15
    smallT.leading = 18
    smallT.leftIndent = -50
    smallT.spaceBefore = 15
    
    
    T = styles['Title']
    T.fontSize = 20
    story = []
    story.append(Paragraph('Report',T))
    story.append(Paragraph('1. Contamination content (Input)',smallT))
    story.append(Paragraph(str('Reference genome Path : ' + reference),p))
    story.append(Paragraph(str('Rawdata Path : ' + rawfastq1),p))
    story.append(Paragraph(str('Rawdata Path : ' + rawfastq2),p))
    f = open(reference,'r')
    l = f.readline()
    targetSpeciesName  = l.split(' ')[1:]
    story.append(Paragraph(str('Target bacterial : ' + " ".join(targetSpeciesName)),p))
    story.append(Paragraph(str('2. Species Summary (output) : '),smallT))
    
    ex = pd.read_excel('output/Species.xlsx')
    data = []
    data = ex.values.tolist()
    for d in data:
        urlNCBI = "ncbi.nlm.nih.gov/\nnuccore/" + d[2]
        urlTaxonomy = 'ncbi.nlm.nih.gov/Taxonomy/Browser/wwwtax.cgi?name=\n' + d[0]
        d[0] = "\n".join(d[0].split(' '))
        d.append(urlNCBI)
        d.append(urlTaxonomy)
    head = ['species name','num of \nmatched reads','accession ID','NCBI dataset URL','Taxonomy dastaset URL']
    data.insert(0,head)
    data = data[:20]
    t = Table(data,colWidths = [90,60,60,90,230],style=[
        ('GRID',(0,0),(-1,-1),1,colors.grey),
        ('FONTSIZE',(0,0),(-1,-1),8),
        ('ALIGN',(0,0),(2,-1),'CENTER'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE')])
    story.append(t)
    
    img = Image('output/Species.png',width = 400, height = 400)
    story.append(Paragraph(str('3. Species Summary Graph (Top ten)'),smallT))
    story.append(img)
    
    story.append(Paragraph(str('4. Contaminated or not ?'),smallT))
    accession = ""
    con_info = []
    for line in conref:
        f = open(line,'r')
        l = f.readline()
        accession = l.split(' ')[0].split('>')[1]
        targetSpeciesName = " ".join(l.split(' ')[1:]).split('\n')[0]
        con_info.append([accession,targetSpeciesName])
        
    if len(con_info) > 0:
        story.append(Paragraph(str('Yes, it is contaminated by : '),p))
        head = ['Contaminated species accesssion ID','Contaminated species name']
        con_info.insert(0,head)
        t = Table(con_info,colWidths = [180,330],style=[
            ('GRID',(0,0),(-1,-1),1,colors.black),
            ('FONTSIZE',(0,0),(-1,-1),9),
            ('ALIGN',(0,0),(-1,-1),'CENTER'),
            ('VALIGN',(0,0),(-1,-1),'MIDDLE')])
        story.append(t) 
        
        story.append(Paragraph(str('5. Variant results'),smallT))
        read1,avecoverage1,read2,avecoverage2,variantsNum,VerifiedMutationNum,misalign = stat_con(reference,'output/raw1.fasta','output/cleanupReads1.fasta')
        data = [['Index','Value'],
                ['Num of reads (before clean)',read1],
                ['Average coverage (before clean)',avecoverage1],
                ['Num of reads (after clean)',read2],
                ['Average coverage (after clean)',avecoverage2],
                ['Num of variants (GATK call)',variantsNum],
                ['Verified mutations',VerifiedMutationNum],
                ['Misalignment',misalign]]
        t = Table(data,colWidths = [200,200],style=[
            ('GRID',(0,0),(-1,-1),1,colors.black)])
        story.append(t)            
    else:
        story.append(Paragraph(str('No, it is clean data.'),p))
        
        story.append(Paragraph(str('5. Variant results'),smallT))
        read1,avecoverage1,variantsNum,VerifiedMutationNum,misalign = stat_notcon(reference,'output/raw1.fasta')
        data = [['Index','Value'],
                ['Num of reads',read1],
                ['Average coverage',avecoverage1],
                ['Num of variants (GATK call)',variantsNum],
                ['Verified mutations',VerifiedMutationNum],
                ['Misalignment',misalign]]
        t = Table(data,colWidths = [200,200],style=[
            ('GRID',(0,0),(-1,-1),1,colors.black)])
        story.append(t)
    doc.build(story)

if __name__=="__main__":
    helpList = ['reference : input your target bacterial referencce genome',
                'rawdata1 : input fastq format',
                'rawdata2 : input fastq format',
                'ntPaht: input nt database path',
                '-similarity : input the threathold of de contamination, the lower the more accuracy',
                '-flash : descide if use flash to merge pair end reads']
    
    path = os.getcwd()   
    if not os.path.isdir(path + "/output"):
        os.makedirs(path + "/output")
    if not os.path.isdir(path + "/output/extract_result"):
        os.makedirs(path + '/output/extract_result')
    if not os.path.isdir(path + "/output/db"):
        os.makedirs(path + '/output/db')
    if not os.path.isdir(path + "/output/GATK"):
        os.makedirs(path + '/output/GATK')
 
    os.system('touch output/Mutquery.fasta')
    os.system('touch output/RKmer.fasta')
    os.system('touch output/PKmer.fasta')
    
    reference = sys.argv[1]#reference genome
    raw1 = sys.argv[2]#raw data 1
    raw2 = sys.argv[3]#raw data 2
    ntdatabasePath = sys.argv[4]#nt db path
    flashFlag = False
    similar = 95
    for i in range(len(sys.argv) - 1):
        if sys.argv[i].split(' ')[0] == '-h':
            print(helpList)
        if sys.argv[i].split(' ')[0] == '-similarity':
            similar = sys.argv[i+1]
        if sys.argv[i].split(' ')[0] == '-flash':
            flashFlag = bool(sys.argv[i+1])
    print("")
    print("Please make sure the output directory is empty.")
    print("cleanSeq: the parameters are as follows:")
    print("reference:",reference)
    print("raw1:",raw1)
    print("raw2:",raw2)
    print("ntPath:",ntdatabasePath)
    print("similarity:",similar)
    print("flashFlag:",flashFlag)
    print("")
    print("First step: quality control.")
    print("")
    rawfastq1 = 'output/raw1.fastq'
    trim1 = 'output/out.trim1.fastq'
    rawfastq2 = 'output/raw2.fastq'
    trim2 = 'output/out.trim2.fastq'
    QCcommand = "java -jar trimmomatic.jar PE -threads 32 -trimlog output/log " + raw1 + ' ' + raw2 + ' '
    QCcommand = QCcommand + rawfastq1 + ' ' + trim1 + ' ' + rawfastq2 + ' ' + trim2 + ' '
    QCcommand = QCcommand + "ILLUMINACLIP:TruSeq3-PE.fa:2:30:10 SLIDINGWINDOW:5:20 LEADING:20 TRAILING:20 MINLEN:35"
    os.system(QCcommand)
    
    fastqtofasta(rawfastq1,'output/raw1.fasta')
    fastqtofasta(rawfastq2,'output/raw2.fasta')
    
    #generate contamination reference list
    print("")
    print("Second step: contamination detection.")
    print("")
    conref = ['1']
    conref = contamiateDetection('output/raw1.fasta',ntdatabasePath,reference)
    #generate cleanReads1 and cleanReads2
    print("Third step: contamination remove.")
    
    if len(conref) > 0:
        print("find contamination, excute cleanup.")
        print("")

        flag = True
        if len(conref) > 1: 
            for con in conref:
                if flag == True:                    
                    cleanup(reference, con, 'output/raw1.fasta', 'output/raw2.fasta',similar)
                    flag = False
                else:
                    os.system("mv output/cleanupReads1.fasta rawdata")
                    os.system("mv output/cleanupReads2.fasta rawdata")
                    cleanup(reference, con, "rawdata/cleanupReads1.fasta", "rawdata/cleanupReads2.fasta",similar)
        else:
            cleanup(reference, conref[0], 'output/raw1.fasta', 'output/raw2.fasta',similar)     
        
        fastaTofastq('output/cleanupReads1.fasta', rawfastq1, 'output/cleanup1.mapp', 'output/cleanupReads1.fastq')
        fastaTofastq('output/cleanupReads2.fasta', rawfastq2, 'output/cleanup2.mapp', 'output/cleanupReads2.fastq')

        grawfastq1 = 'output/cleanupReads1.fastq'
        grawfastq2 = 'output/cleanupReads2.fastq'
        grawfasta1= 'output/cleanupReads1.fasta'
        grawfasta2 = 'output/cleanupReads2.fasta'
    else:
        print("no contamination found.")
        print("")
        grawfastq1 = 'output/raw1.fastq'
        grawfastq2 = 'output/raw2.fastq'
        grawfasta1 = 'output/raw1.fasta'
        grawfasta2 = 'output/raw2.fasta'
    
    #call variants
    print("Fourth step: mutation calling.")
    print("")
    gref = reference.split('.')[0]
    gatkCommand = './GATK.sh '+ grawfastq1 + ' ' + grawfastq2 + ' ' + gref + ' output/GATK/R'
    os.system(gatkCommand)
    print("")
    print("Fifth step: mutation verification.")
    print("")
    #mutationDetection
    if flashFlag:
        command = './flash ' + grawfastq1 + ' ' + grawfastq2 + ' -o output/clean >> output/log'
        os.system(command)
        fastqtofasta('output/clean.extendedFrags.fastq', 'output/clean.extendedFrags.fasta')
        singleEndBlast(reference, 'output/GATK/R.filter.vcf', 'output/clean.extendedFrags.fasta')
    else:
        pairEndBlast(reference,'output/GATK/R.filter.vcf', grawfasta1, grawfasta2)
    #write report
    print("")
    print("Sixth step: generate report.")
    print("")
    contaminationReport(reference,raw1,raw2,conref)
    delet()
    print("All done.")
    